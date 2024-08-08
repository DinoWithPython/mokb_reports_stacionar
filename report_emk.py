import datetime
from copy import deepcopy
import os
import sys
from typing import Any, Optional, Tuple, Union

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Color, Font, PatternFill
from openpyxl.styles.borders import Border, Side

from validators import (
    validate_column_with_data,
    validate_for_title,
    validate_numbers,
    validate_not_pdo,
    ValidateError)

if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
elif __file__:
    BASE_DIR = os.path.dirname(__file__)

NUMBER_OF_INDICATORS = 16
BEGIN_INDICATORS_IN_ROW = 9
EXPECTED_MIN_COLUMN_VALUES = 21
TITLE_VALUES = ["Наименование Медицинской организации", ]
PDO_NAMES = ["приемн", "приёмн", " ПДО "]
HEADINGS = {
    0: "Номер КВС",
    1: "Дата выписки из стац",
    2: "Дата и время выписки из указанного движения",
    3: "Возраст на момент госпитализации в стационар",
    4: "Наличие ПОЛИСА",
    5: "Наличие ДУЛ",
    6: "Наличие СНИЛС",
    7: "Гражданство",
    8: "Отделение",
    9: "Наличие заполненного первичного осмотра  в указанном движении",
    10: "Количество дневниковых записей, которое необходимо было завести в указанном движении",
    11: "Количество оформленных дневниковых записей  в указанном движении",
    # 12: "Наличие оформленного эпикриза в указанном движении",
    12: "Эпикриз подписан ЭЦП в указанном движении",
    13: "Хир активность (операции)",
    14: "Хир активность (количество)",
    15: "Хир активность (протоколы)",
    16: "Наличие оформленных лекарственных назначений в указанном движении",
    17: "Количество оформленных направлений на лабораторные исследования в указанном движении",
    18: "Количество проведенных лабораторных исследований в указанном движении",
    19: "Количество оформленных направлений на инструментальные методы лечения в указанном движении",
    20: "Количество проведенных инструментальных исследований  в указанном движении",
    21: "Количество оформленных направлений на консультативные услуги в указанном движении",
    22: "Количество оформленных  консультативных услуг в указанном движении",
    23: "Количество необходимых реанимационных дневников в указанном движении",
    24: "Количество оформленных реанимационных дневников в указанном движении",
}


def now():
    """Возвращает текущую дату и время в формате строки."""
    return datetime.datetime.now().strftime("%d.%m %H_%M_%S")


def log_any_error(any_text):
    """Функция для записи ошибок в файл."""
    log = open(rf"{BASE_DIR}\log_any_error.txt", "a+", encoding='UTF-8')
    text = str(any_text)
    print(f"[{now()}] {text}", file=log)
    log.close()


def check_index(lst: list, column_name: str, class_name: str = None):
    """Получает название столбца и возвращает индекс."""
    try:
        upper_list = [element.upper() if isinstance(element, str) else element for element in lst]
        index_el = upper_list.index(column_name.upper())
        return index_el
    except ValueError:
        if class_name:
            log_any_error(f'Не найден столбец "{column_name}" в классе {class_name}')
            return
        log_any_error(f'Не найден столбец "{column_name}"')


def procent_is_none(lst: list) -> str:
    """Если нет процента в исходном файле, то считаем сами."""
    indicators = [0 if x is None else 1 for x in lst[BEGIN_INDICATORS_IN_ROW:]]
    procent = ((sum(indicators) - 1) / NUMBER_OF_INDICATORS) * 100
    procent = int(procent + (0.5 if procent > 0 else -0.5))
    return str(procent)


def date_conversion(date: Union[str, datetime.datetime]):
    """Преобразует дату к нужному формату."""
    if isinstance(date, datetime.datetime):
        return date.strftime("%d.%m.%Y")
    if isinstance(date, str):
        raise ValueError("Попытка преобразовать дату из строки, а не datetime. Автор не был к такому готов. Нужна переработка скрипта. :)")


def check_convert_type(variable: Any, type_variable: Any, name_variable: str):
    """Проверяет возможность преобразования в переданный тип."""
    try:
        type_variable(variable)
    except ValueError as e:
        text_err = (
            f"Ошибка в типе переменной. Типом '{variable}' является '{type(variable)}'."
            f"\nОжидается преобразование {name_variable} в '{type_variable}'!")
        log_any_error(f'{text_err}, {e}')
        raise ValueError(text_err)


class EmkDataFromFile:
    """Класс для хранения данных под вложенные в ЭМК отчеты."""
    data = None
    title = None


class EmkReport:
    """Класс для создания отчетов по ЭМК. Принимает путь для файла."""
    period = set()

    def __init__(self, filepath: str, need_pdo: bool = False):
        self.filepath = filepath
        self.need_pdo = need_pdo
        self.kvs_number = None
        self.date_out_from_hospital = None
        self.date_out_from_stage = None
        self.age = None
        self.have_polis = None
        self.have_dul = None
        self.have_snils = None
        self.citizenship = None
        self.department = None
        self.is_initial_exam = None
        self.count_diary_needed = None
        self.count_diary = None
        # self.is_epicrisis = None
        self.epicrisis_with_ecp = None
        self.surgical_operation_name = None
        self.surgical_operation_count = None
        self.surgical_operation_protocols = None
        self.is_medicinal_purposes = None
        self.count_lab_research = None
        self.count_lab_research_complete = None
        self.count_inst_research = None
        self.count_inst_research_complete = None
        self.count_cons = None
        self.count_cons_complete = None
        self.count_needed_rean = None
        self.count_input_rean = None

    def __str__(self):
        return "EmkNewReport"

    def validate_data_from_file(self, works_sheet, for_data: list, title_excel: list):
        """Валидация данных из файла."""
        for row in works_sheet.iter_rows(values_only=True):
            if 'lpu_name' in row:
                continue
            if validate_column_with_data(row, EXPECTED_MIN_COLUMN_VALUES):
                continue
            if validate_numbers(row):
                continue
            if len(title_excel) == 0:
                if validate_for_title(TITLE_VALUES, row):
                    title_excel.extend(row)
                    self.kvs_number = check_index(title_excel, HEADINGS[0], self)
                    self.date_out_from_hospital = check_index(title_excel, HEADINGS[1], self)
                    self.date_out_from_stage = check_index(title_excel, HEADINGS[2], self)
                    self.age = check_index(title_excel, HEADINGS[3], self)
                    self.have_polis = check_index(title_excel, HEADINGS[4], self)
                    self.have_dul = check_index(title_excel, HEADINGS[5], self)
                    self.have_snils = check_index(title_excel, HEADINGS[6], self)
                    self.citizenship = check_index(title_excel, HEADINGS[7], self)
                    self.department = check_index(title_excel, HEADINGS[8], self)
                    self.is_initial_exam = check_index(title_excel, HEADINGS[9], self)
                    self.count_diary_needed = check_index(title_excel, HEADINGS[10], self)
                    self.count_diary = check_index(title_excel, HEADINGS[11], self)
                    # self.is_epicrisis = check_index(title_excel, HEADINGS[12], self)
                    self.epicrisis_with_ecp = check_index(title_excel, HEADINGS[12], self)
                    self.surgical_operation_name = check_index(title_excel, HEADINGS[13], self)
                    self.surgical_operation_count = check_index(title_excel, HEADINGS[14], self)
                    self.surgical_operation_protocols = check_index(title_excel, HEADINGS[15], self)
                    self.is_medicinal_purposes = check_index(title_excel, HEADINGS[16], self)
                    self.count_lab_research = check_index(title_excel, HEADINGS[17], self)
                    self.count_lab_research_complete = check_index(title_excel, HEADINGS[18], self)
                    self.count_inst_research = check_index(title_excel, HEADINGS[19], self)
                    self.count_inst_research_complete = check_index(title_excel, HEADINGS[20], self)
                    self.count_cons = check_index(title_excel, HEADINGS[21], self)
                    self.count_cons_complete = check_index(title_excel, HEADINGS[22], self)
                    self.count_needed_rean = check_index(title_excel, HEADINGS[23], self)
                    self.count_input_rean = check_index(title_excel, HEADINGS[24], self)
                    if None in self.__dict__.values():
                        raise ValueError("Отсутвует необходимый столбец. Подробности в файле log_any_error.txt")
                    continue
            if self.need_pdo:
                for_data.append(row)
                continue
            if validate_not_pdo(PDO_NAMES, row[self.department]):
                for_data.append(row)

        if len(title_excel) == 0:
            log_any_error(f"Заголовки не были найдены! Ни в одной строке не было: \n{TITLE_VALUES}\n")
            raise ValueError("В файле отсутствуют заголовки! Подробности в файле log_any_error.txt")
        if len(for_data) == 0:
            log_any_error("Файл с данными пуст!")
            raise ValueError("Файл с данными пуст! Проверьте, что выбрали нужный файл.")

    def open_file_return_data(self) -> list:
        """Открыли файл и вернули истину и данные, либо ложь и ошибку."""
        try:
            wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
            ws = wb.active

            data_from_excel = []
            title_excel = []
            self.validate_data_from_file(ws, data_from_excel, title_excel)
            EmkDataFromFile()
            EmkDataFromFile.data = data_from_excel
            EmkDataFromFile.title = title_excel
            return data_from_excel
        except TypeError as e:
            raise ValidateError(e)
        except Exception as e:
            raise ValidateError(e)
        finally:
            try:
                wb._archive.close()
            except Exception as e:
                log_any_error(f"[ERR] Ошибка при закрытии файла. \n{e}")

    def operation_with_data(self, record: list, data_summary: dict) -> list:
        """Обрабатывает список и наполняет словарь."""
        temp = [0] * len(HEADINGS)
        temp[0] = record[self.kvs_number]
        temp[1] = record[self.date_out_from_hospital]
        self.period.add(temp[1])
        temp[2] = record[self.date_out_from_stage]
        temp[3] = record[self.age]
        temp[4] = record[self.have_polis]
        temp[5] = record[self.have_dul]
        temp[6] = record[self.have_snils]
        temp[7] = record[self.citizenship]
        temp[8] = record[self.department]
        if "Нет" in (temp[4], temp[5], temp[6]):
            if temp[8] not in data_summary["Не указаны перс.данные"]:
                data_summary["Не указаны перс.данные"][temp[8]] = {temp[0]: 1}
            else:
                if (temp[0] not in data_summary["Не указаны перс.данные"][temp[8]]):
                    data_summary["Не указаны перс.данные"][temp[8]][temp[0]] = 1
                else:
                    data_summary["Не указаны перс.данные"][temp[8]][temp[0]] += 1
        temp[9] = record[self.is_initial_exam]
        if temp[9] is None:
            if temp[8] not in data_summary["Нет первичного осмотра"]:
                data_summary["Нет первичного осмотра"][temp[8]] = {temp[0]: 1}
            else:
                if temp[0] not in data_summary["Нет первичного осмотра"][temp[8]]:
                    data_summary["Нет первичного осмотра"][temp[8]][temp[0]] = 1
                else:
                    data_summary["Нет первичного осмотра"][temp[8]][temp[0]] += 1
        temp[10] = record[self.count_diary_needed]
        temp[11] = record[self.count_diary]
        if temp[11] is None:
            if temp[8] not in data_summary["Нет оформленных дневниковых записей"]:
                data_summary["Нет оформленных дневниковых записей"][temp[8]] = {
                    temp[0]: 1
                }
            else:
                if (
                    temp[0]
                    not in data_summary["Нет оформленных дневниковых записей"][
                        temp[8]
                    ]
                ):
                    data_summary["Нет оформленных дневниковых записей"][temp[8]][
                        temp[0]
                    ] = 1
                else:
                    data_summary["Нет оформленных дневниковых записей"][temp[8]][
                        temp[0]
                    ] += 1
        # temp[12] = record[self.is_epicrisis]
        # if temp[12] is None:
        #     if temp[8] not in data_summary["Нет оформленных эпикризов"]:
        #         data_summary["Нет оформленных эпикризов"][temp[8]] = {temp[0]: 1}
        #     else:
        #         if (
        #             temp[0]
        #             not in data_summary["Нет оформленных эпикризов"][temp[8]]
        #         ):
        #             data_summary["Нет оформленных эпикризов"][temp[8]][temp[0]] = 1
        #         else:
        #             data_summary["Нет оформленных эпикризов"][temp[8]][temp[0]] += 1

        # Далее все заголовки были снижены на 1 после 12.
        temp[12] = record[self.epicrisis_with_ecp]
        if temp[12] is None or temp[12].upper() == "нет".upper():
            if temp[8] not in data_summary["Выписной эпикриз не подписан ЭЦП"]:
                data_summary["Выписной эпикриз не подписан ЭЦП"][temp[8]] = {
                    temp[0]: 1
                }
            else:
                if (
                    temp[0]
                    not in data_summary["Выписной эпикриз не подписан ЭЦП"][temp[8]]
                ):
                    data_summary["Выписной эпикриз не подписан ЭЦП"][temp[8]][
                        temp[0]
                    ] = 1
                else:
                    data_summary["Выписной эпикриз не подписан ЭЦП"][temp[8]][
                        temp[0]
                    ] += 1
        temp[13] = record[self.surgical_operation_name]
        temp[14] = record[self.surgical_operation_count]
        temp[15] = record[self.surgical_operation_protocols]
        temp[16] = record[self.is_medicinal_purposes]
        if temp[16] is None:
            if (
                temp[8]
                not in data_summary["Нет назначений лекарственных препаратов"]
            ):
                data_summary["Нет назначений лекарственных препаратов"][temp[8]] = {
                    temp[0]: 1
                }
            else:
                if (
                    temp[0]
                    not in data_summary["Нет назначений лекарственных препаратов"][
                        temp[8]
                    ]
                ):
                    data_summary["Нет назначений лекарственных препаратов"][
                        temp[8]
                    ][temp[0]] = 1
                else:
                    data_summary["Нет назначений лекарственных препаратов"][
                        temp[8]
                    ][temp[0]] += 1
        temp[17] = record[self.count_lab_research]
        temp[18] = record[self.count_lab_research_complete]
        temp[19] = record[self.count_inst_research]
        temp[20] = record[self.count_inst_research_complete]
        temp[21] = record[self.count_cons]
        temp[22] = record[self.count_cons_complete]
        temp[23] = record[self.count_needed_rean]
        temp[24] = record[self.count_input_rean]
        return temp

    def svod_on_sheet(self, sheet, data_summary: dict, lst: list):
        """Добавляет данные и оформление для листа свода."""
        for key1, value1 in data_summary.items():
            if len(value1) > 0:
                lst.append([key1])
                count_all = 0
                for key2, value2 in value1.items():
                    if len(value2) > 0:
                        count = sum(value2.values())
                        lst.append([key2, count])
                        count_all += count
                        for key3, value3 in value2.items():
                            lst.append([f"        {key3}", value3])
                lst.append(["Итого:", count_all])
                lst.append(None)

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        my_color = Color(rgb="AFEEEE")
        my_fill = PatternFill(patternType="solid", fgColor=my_color)
        column = 0
        row = 0
        if len(self.period) > 1:
            date_min = date_conversion(min(self.period))
            date_max = date_conversion(max(self.period))
        else:
            date_min = date_max = date_conversion(self.period)
        for r in range(0, len(lst)):
            if lst[r] is None:
                column += 2
                row = 0
                continue
            if len(lst[r]) > 1:
                sheet.cell(row=row + 1, column=column + 1).value = lst[r][0]
                sheet.cell(row=row + 1, column=column + 1).border = thin_border
                sheet.cell(row=row + 1, column=column + 2).value = lst[r][1]
                sheet.cell(row=row + 1, column=column + 2).border = thin_border
                row += 1
                continue
            sheet.cell(row=row + 1, column=column + 1).value = lst[r][0]
            sheet.cell(row=row + 1, column=column + 1).border = thin_border
            sheet.cell(row=row + 1, column=column + 1).font = Font(bold=True)
            sheet.cell(row=row + 1, column=column + 1).fill = my_fill
            if len(self.period) > 1:
                sheet.cell(row=row + 1, column=column + 2).value = f"{date_min}-{date_max}"
                sheet.cell(row=row + 1, column=column + 2).border = thin_border
            else:
                sheet.cell(row=row + 1, column=column + 2).value = f"{date_min}"
                sheet.cell(row=row + 1, column=column + 2).border = thin_border
            # sheet.merge_cells(
            #     start_row=1, start_column=column+1, end_row=1, end_column=column+2
            # )
            row += 1

        # Задаю ширину столбцев свода
        width_column_svod = 80
        sheet.column_dimensions["A"].width = width_column_svod
        sheet.column_dimensions["C"].width = width_column_svod
        sheet.column_dimensions["E"].width = width_column_svod
        sheet.column_dimensions["G"].width = width_column_svod
        sheet.column_dimensions["I"].width = width_column_svod
        sheet.column_dimensions["K"].width = width_column_svod

    def personal_on_sheet(self, lst: list, sheet):
        """Для листа персонально."""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Для высчитывания процента
        for element in lst:
            # if element[8] is None:
            #     element[8] = procent_is_none(element)
            sheet.append(element)
        sheet.auto_filter.ref = sheet.dimensions

        red_color = Color(rgb="FF0000")
        red_fill = PatternFill(patternType="solid", fgColor=red_color)
        green_color = Color(rgb="32CD32")
        green_fill = PatternFill(patternType="solid", fgColor=green_color)
        low_procent = Color(rgb="CD5C5C")
        # low_fill = PatternFill(patternType="solid", fgColor=low_procent)
        hign_procent = Color(rgb="F08080")
        # hign_fill = PatternFill(patternType="solid", fgColor=hign_procent)
        for r in sheet.iter_rows(max_row=1):
            for column in r:
                column.border = thin_border
                column.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
        indexes = {value: key for key, value in HEADINGS.items()}
        for r in sheet.iter_rows(min_row=2):
            date_out_from_hospital = date_conversion(r[
                indexes["Дата выписки из стац"]
                ].value)
            r[indexes["Дата выписки из стац"]].value = date_out_from_hospital
            date_out_from_stage = date_conversion(r[
                indexes["Дата и время выписки из указанного движения"]
                ].value)
            r[
                indexes["Дата и время выписки из указанного движения"]
                ].value = date_out_from_stage

            department = r[indexes["Отделение"]]
            initial_exam = r[indexes["Наличие заполненного первичного осмотра  в указанном движении"]]
            initial_exam.fill = red_fill if initial_exam.value is None else green_fill
            count_dairy_need = r[indexes["Количество дневниковых записей, которое необходимо было завести в указанном движении"]]
            count_dairy = r[indexes["Количество оформленных дневниковых записей  в указанном движении"]]
            check_convert_type(count_dairy.value, int, "count_dairy")
            check_convert_type(count_dairy_need.value if count_dairy_need.value is not None else 0, int, "count_dairy_need")
            if int(count_dairy.value) < int(count_dairy_need.value if count_dairy_need.value is not None else 0):
                count_dairy.fill = red_fill
            else:
                count_dairy.fill = green_fill
            # epi = r[indexes["Наличие оформленного эпикриза в указанном движении"]]
            epi_ecp = r[indexes["Эпикриз подписан ЭЦП в указанном движении"]]
            # epi.fill = red_fill if epi.value is None else green_fill
            epi_ecp.fill = red_fill if epi_ecp.value in (None, "Нет") else green_fill
            have_polis = r[indexes["Наличие ПОЛИСА"]]
            have_dul = r[indexes["Наличие ДУЛ"]]
            have_snils = r[indexes["Наличие СНИЛС"]]
            have_polis.fill = red_fill if have_polis.value == "Нет" else green_fill
            have_dul.fill = red_fill if have_dul.value == "Нет" else green_fill
            have_snils.fill = red_fill if have_snils.value == "Нет" else green_fill
            count_operation = r[indexes["Хир активность (количество)"]]
            count_input_operation = r[indexes["Хир активность (протоколы)"]]
            check_convert_type(count_input_operation.value, int, "count_input_operation")
            check_convert_type(count_operation.value, int, "count_operation")
            if int(count_input_operation.value) < int(count_operation.value):
                count_input_operation.fill = red_fill
            else:
                count_input_operation.fill = green_fill
            count_consultation = r[indexes['Количество оформленных направлений на консультативные услуги в указанном движении']]
            count_input_consultation = r[indexes['Количество оформленных  консультативных услуг в указанном движении']]
            check_convert_type(count_consultation.value, int, "count_consultation")
            check_convert_type(count_input_consultation.value, int, "count_input_consultation")
            if int(count_input_consultation.value) < int(count_consultation.value):
                count_input_consultation.fill = red_fill
            count_reanim = r[indexes["Количество необходимых реанимационных дневников в указанном движении"]]
            count_input_reanim = r[indexes["Количество оформленных реанимационных дневников в указанном движении"]]
            check_convert_type(count_input_reanim.value, int, "count_input_reanim")
            check_convert_type(count_reanim.value, int, "count_reanim")
            if int(count_input_reanim.value) < int(count_reanim.value):
                count_input_reanim.fill = red_fill
            else:
                count_input_reanim.fill = green_fill
            for cell in r:
                cell.border = thin_border
            department.alignment = Alignment(wrap_text=True)

        sheet.column_dimensions["A"].width = 15
        sheet.column_dimensions["B"].width = 15
        sheet.column_dimensions["C"].width = 15
        sheet.column_dimensions["I"].width = 30
        columns_for_dimensions = [
            "J",
            "K",
            "L",
            "M",
            "N",
            "O",
            "P",
            "Q",
            "R",
            "S",
            "T",
            "U",
            "V",
            "W",
            "X",
            "Y",
            "Z",
            "AA",
        ]
        for letter in columns_for_dimensions:
            sheet.column_dimensions[letter].width = 18

    def processing_report(
        self,
        data: list,
        selected_day: str = None) -> Tuple[list, list, list, list]:
        """Обрабатывает все или конкретную дату."""
        data_summary_kc = {
            "Не указаны перс.данные": {},
            "Нет первичного осмотра": {},
            "Нет оформленных дневниковых записей": {},
            "Нет назначений лекарственных препаратов": {},
            # "Нет оформленных эпикризов": {},
            "Выписной эпикриз не подписан ЭЦП": {},
        }

        data_summary_dc = deepcopy(data_summary_kc)

        data_personal_kc = [
            [element for element in HEADINGS.values()],
        ]
        data_personal_dc = deepcopy(data_personal_kc)

        if selected_day is None:
            for record in data:
                if (
                    "ДС" in record[self.department]
                    or "дневн" in record[self.department].lower()
                    or "дн.стац." in record[self.department].lower()
                    or "дн. стац." in record[self.department].lower()
                ):
                    temp = self.operation_with_data(record, data_summary_dc)
                    data_personal_dc.append(temp)
                    continue
                temp = self.operation_with_data(record, data_summary_kc)
                data_personal_kc.append(temp)
        else:
            for record in data:
                if date_conversion(record[self.date_out_from_hospital]) != selected_day:
                    continue
                if (
                    "ДС" in record[self.department]
                    or "дневн" in record[self.department].lower()
                    or "дн.стац." in record[self.department].lower()
                    or "дн. стац." in record[self.department].lower()
                ):
                    temp = self.operation_with_data(record, data_summary_dc)
                    data_personal_dc.append(temp)
                    continue
                temp = self.operation_with_data(record, data_summary_kc)
                data_personal_kc.append(temp)
        return (data_summary_kc, data_summary_dc, data_personal_kc, data_personal_dc)

    def save_files(self, data_summary_kc, data_summary_dc, data_personal_kc, data_personal_dc):
        """Функция для сохранения файлов."""
        wb = Workbook()
        wb.create_sheet("Наполнение КВС")
        wb.create_sheet("Свод ДС")
        wb.create_sheet("Наполнение КВС ДС")

        sheet_svod_kc = wb.active
        sheet_svod_kc.title = "Свод"
        svod_kc = []
        self.svod_on_sheet(sheet_svod_kc, data_summary_kc, svod_kc)
        sheet_svod_dc = wb["Свод ДС"]
        svod_dc = []
        self.svod_on_sheet(sheet_svod_dc, data_summary_dc, svod_dc)

        sheet_personal_kc = wb["Наполнение КВС"]
        self.personal_on_sheet(data_personal_kc, sheet_personal_kc)
        sheet_personal_dc = wb["Наполнение КВС ДС"]
        self.personal_on_sheet(data_personal_dc, sheet_personal_dc)

        wb.save(f"Свод на {now()}.xlsx")


class CheckIdentificator:
    """Универсальный класс для формирования таблички переданных индикаторов."""
    def __init__(self, indicator: str = None, check_indicator: str = None) -> None:
        self.indicator: str = indicator
        self.check_indicator: str = check_indicator
        self.date_out: int = None
        self.department: int = None
        self.indicator_index: int = None
        self.check_indicator_index: int = None
        self.period = set()

    def __str__(self):
        return self.__class__.__name__

    def title_indicator(self):
        """
        Необходимо определить метод, что вернет список заголовков индикатора.
        title_lis = [
            "Подразделение",
            "Количество движений",
            "Наличие оформленных направлений на лабораторные исследования",
            "Процент направлений",
            "НЕТ оформленных направлений на лабораторные исследования",
            "Отметки о проведении лабораторных исследований (Да)",
            "Процент выполнения от созданных",
            "НЕТ Отметки о проведении лабораторных исследований (Нет)",
        ]
        """
        pass

    def validate_date(self):
        """Присваиваем индексы и прочее."""
        EmkDataFromFile()
        title = EmkDataFromFile.title
        self.date_out = check_index(title, HEADINGS[1], self)
        self.department = check_index(title, HEADINGS[8], self)
        self.indicator_index = check_index(title, self.indicator, self)
        self.check_indicator_index = check_index(title, self.check_indicator, self)

    def processing(self):
        """Обработка данных и формирование листов с индикатором."""
        self.validate_date()
        researches = {}
        data = EmkDataFromFile.data
        COLUMNS = {
            "mark_direction": 0,
            "mark_complete_direction": 1,
            "procent_direction": 2,
        }
        for row in data:
            if row[self.department] not in researches:
                researches[row[self.department]] = [0] * len(COLUMNS)
            self.period.add(row[self.date_out])
            if row[self.indicator_index] != 0:
                researches[row[self.department]][COLUMNS["mark_direction"]] += row[self.indicator_index]
            if row[self.check_indicator_index] != 0:
                researches[row[self.department]][COLUMNS["mark_complete_direction"]] += row[self.check_indicator_index]

        results = [0] * len(COLUMNS)
        for key in researches:
            researches[key][COLUMNS["procent_direction"]] = (
                (
                    researches[key][COLUMNS["mark_complete_direction"]]
                    / researches[key][COLUMNS["mark_direction"]]
                )
                if researches[key][COLUMNS["mark_direction"]] != 0
                else 0
            )
            for element in COLUMNS.values():
                results[element] += researches[key][element]

        results[COLUMNS["procent_direction"]] = (
            (
                results[COLUMNS["mark_complete_direction"]]
                / results[COLUMNS["mark_direction"]]
            )
            if results[COLUMNS["mark_direction"]] != 0
            else 0
        )

        researches["Итого"] = results
        return researches

    def data_in_sheet(self, sheet, header: list, title: list, lst: list):
        """Добавляет данные на лист."""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        sheet.append(header)
        sheet.append(title)
        for row in lst:
            sheet.append(row)

        sheet.auto_filter.ref = "A2:D2"
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        my_color = Color(rgb="AFEEEE")
        my_fill = PatternFill(patternType="solid", fgColor=my_color)
        for r in sheet.iter_rows(max_row=2):
            for column in r:
                column.border = thin_border
                column.fill = my_fill
                column.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

        red_color = Color(rgb="FF0000")
        red_fill = PatternFill(patternType="solid", fgColor=red_color)
        green_color = Color(rgb="32CD32")
        green_fill = PatternFill(patternType="solid", fgColor=green_color)
        for r in sheet.iter_rows(min_row=3):
            for column in range(len(r)):
                r[column].border = thin_border
                r[column].alignment = Alignment(horizontal="center", vertical="center")
                if column == 1 or column == 2:
                    if r[column].value == 0:
                        r[column].value = None
                if column == 3:
                    r[column].fill = (
                        red_fill if float(r[column].value) < 0.90 else green_fill
                    )
                    r[column].number_format = "0%"

        sheet.column_dimensions["A"].width = 100
        columns_for_dimensions = ["B", "C", "D"]
        for letter in columns_for_dimensions:
            sheet.column_dimensions[letter].width = 15

    def save_file(
        self,
        data: dict,
        name_sheet: str,
        header_sheet: str,
        file_name: str):
        """Сохраняет файл необходимого индикатора."""
        to_excel = []
        if len(self.period) > 1:
            date_min = date_conversion(min(self.period))
            dete_max = date_conversion(max(self.period))
        else:
            date_min = dete_max = date_conversion(min(self.period))

        for key, value in data.items():
            to_excel.append([key] + value)
        
        wb_indicator = Workbook()
        sheet_indicator = wb_indicator.active
        sheet_indicator.title = name_sheet
        header_indicator = [f"{header_sheet} с {date_min} по {dete_max}"]
        title_indicator = self.title_indicator()
        self.data_in_sheet(sheet_indicator, header_indicator, title_indicator, to_excel)
        wb_indicator.save(f"{file_name} на {now()}.xlsx")


class LisIdentificator(CheckIdentificator):
    """Обрабатывает и сохраняет отчет по ЛИС."""
    def __init__(self, indicator: str = None, check_indicator: str = None) -> None:
        super().__init__(indicator, check_indicator)

    def title_indicator(self):
        return [
            "Подразделение",
            "Количество направлений на лабораторные исследования",
            "Количество проведенных лабораторных исследований",
            "Процент выполненных направлений на лабораторные исследования",
            ]


class InstIdentificator(CheckIdentificator):
    """Обрабатывает и сохраняет отчет по Инструментальным исследованиям."""
    def __init__(self, indicator: str = None, check_indicator: str = None) -> None:
        super().__init__(indicator, check_indicator)

    def title_indicator(self):
        return [
            "Подразделение",
            "Количество оформленных направлений на инструментальные методы лечения",
            "Количество проведенных инструментальных исследований",
            "Процент выполненных направлений на инструментальные исследования",
            ]


class ConsIdentificator(CheckIdentificator):
    """Обрабатывает и сохраняет отчет по Консультациям."""
    def __init__(self, indicator: str = None, check_indicator: str = None) -> None:
        super().__init__(indicator, check_indicator)

    def title_indicator(self):
        return [
            "Подразделение",
            "Количество оформленных направлений на консультации",
            "Количество проведенных консультаций",
            "Процент выполненных направлений на консультации",
            ]



if __name__ == "__main__":

    def test_1():
        filepath = rf"{BASE_DIR}/Отчет по ЭМК.xlsx"

        emk = EmkReport(filepath)
        data = emk.open_file_return_data()
        emk.all_days(data[1], data[2])
        # emk.all_days(data[1], data[2], '27.02.2023')

    # test_1()

    def test_3():
        filepath = rf"{BASE_DIR}/ЭМК_стац_20.05.2024-26.05.2024.xlsx"
        emk = EmkReport(filepath)
        data = emk.open_file_return_data()
        for_saves = emk.processing_report(data)
        emk.save_files(*for_saves)
    
    test_3()

    def test_4():
        filepath = rf"{BASE_DIR}/ЭМК_стац_26.06.2023-02.07.2023.xlsx"
        emk = EmkReport(filepath, True)
        data_emk = emk.open_file_return_data()
        data = emk.processing_report(data_emk)
        emk.save_files(*data)


        # lis = LisIdentificator(HEADINGS[18], HEADINGS[19])
        # data = lis.processing()
        # lis.save_file(data,"Отчет по ЛИС в ЭМК", "ЛИС по выписанным пациентам", "ЛИС")
        # ins = InstIdentificator(HEADINGS[20], HEADINGS[21])
        # data_ins = ins.processing()
        # ins.save_file(data_ins, "Отчет по Инстр.напр. в ЭМК", "Инструментальная диагностика по выписанным пациентам", "Инструм на")
        # cons = ConsIdentificator(HEADINGS[22], HEADINGS[23])
        # data_cons = cons.processing()
        # cons.save_file(data_cons, "Отчет по Конс. в ЭМК", "Оформление консультативных услуг по выписанным пациентам", "Консультации на")
    # test_4()
