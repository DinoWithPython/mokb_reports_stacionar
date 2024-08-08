# -*- coding: utf-8 -*-
import datetime
import os
import sys
from typing import Optional, Tuple, Union

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Color, Font, PatternFill
from openpyxl.styles.borders import Border, Side

from validators import (
    validate_department,
    validate_count_days,
    validate_number_history,
    validate_bunks_from_file,
    validate_numbers,
    validate_column_with_data,
    validate_for_title,
    ValidateError
)

if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
elif __file__:
    BASE_DIR = os.path.dirname(__file__)


COUNT_DAYS = 50
EXPECTED_MIN_COLUMN_VALUES = 6
BUNKS = {}
TITLE_VALUES = ["ФИО пациента",]


def now() -> str:
    """Возвращает текущую дату и время в формате строки."""
    return datetime.datetime.now().strftime("%d.%m %H_%M_%S")


def log_any_error(any_text):
    """Функция для записи ошибок в файл."""
    log = open(rf"{BASE_DIR}\log_any_error.txt", "a+", encoding='UTF-8')
    text = str(any_text)
    print(f"[{now()}] {text}", file=log)
    log.close()


def check_index(lst: list, column_name: str, class_name: str = None):
    """Получает название столбца и возвращает индекс."""
    try:
        upper_list = [element.upper() if isinstance(element, str) else element for element in lst]
        index_el = upper_list.index(column_name.upper())
        return index_el
    except ValueError:
        if class_name:
            log_any_error(f'Не найден столбец "{column_name}" в классе {class_name}')
            return
        log_any_error(f'Не найден столбец "{column_name}"')


class BunkReport:
    """Класс для создания отчетов по ЭМК. Принимает путь для файла."""

    def __init__(self, filepath: str):
        self.filepath = filepath

    def open_file_with_bunks(self) -> bool:
        """Фукнция для открытия файла с койками для использования далее."""
        global BUNKS

        file_name = "Отделения и койки.xlsx"
        file_exists = os.path.exists(rf"{BASE_DIR}/{file_name}")
        if file_exists:
            wb = openpyxl.load_workbook(
                rf"{BASE_DIR}/{file_name}", read_only=True, data_only=True
            )
            sheet = wb.active

            if len(BUNKS) > 0:
                BUNKS = {}
            for row in sheet.iter_rows(min_row=2, values_only=True):
                department = row[0]
                validate_bunks_from_file(row[1], department, BUNKS)
            wb._archive.close()
            return True
        self.create_sample()
        return False

    def validate_data_for_filepath(self, works_sheet, data_bunks: dict, data_50: list):
        """Функция валидации строк, при чтении файла."""
        department_column = None
        is_title = True
        title = []
        for row in works_sheet.iter_rows(values_only=True):
            if validate_column_with_data(row, EXPECTED_MIN_COLUMN_VALUES):
                continue
            if validate_numbers(row):
                continue
            if is_title:
                if validate_for_title(TITLE_VALUES, row):
                    is_title = False
                    title = row
                    department_column_ind = check_index(title, "Отделение")
                    if department_column_ind is None:
                        raise ValidateError("Не найден столбец 'Отделение'!")
                    department_column = department_column_ind
                continue
            if data_bunks.get(row[department_column]):
                data_bunks[row[department_column]] += 1
            else:
                if not validate_department(row[department_column]):
                    continue
                data_bunks[row[department_column]] = 1
            count_days_index = check_index(title, "Кол-во \nк/дней")
            if count_days_index is None:
                raise ValidateError("Не найден столбец 'Кол-во \nк/дней'!")
            count_days = validate_count_days(row[count_days_index])
            if count_days >= COUNT_DAYS:
                department = row[department_column]
                number_history_index = check_index(title, "Номер истории болезни")
                if number_history_index is None:
                    raise ValidateError("Не найден столбец 'Номер истории болезни'!")
                number_history = validate_number_history(row[number_history_index])
                data_50.append([department, number_history, count_days])
        if is_title:
            raise ValidateError("Скорее всего Вы выбрали не тот файл или в нём нет заголовков!")

    def open_file_return_data(self) -> Tuple[dict, list]:
        """Открыли файл и вернули истину и данные, либо ложь и ошибку."""
        try:
            wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
            data_bunks = {}
            data_50 = []
            sheets = wb.sheetnames
            if len(sheets) == 1:
                ws = wb.active
                self.validate_data_for_filepath(ws, data_bunks, data_50)
            elif len(sheets) == 0:
                raise ValidateError('В файле отчета нет листов.')
            else:
                for sheet in sheets:
                    ws = wb[sheet]
                    self.validate_data_for_filepath(ws, data_bunks, data_50)
            
            return data_bunks, data_50
        except TypeError as e:
            log_any_error(f"[ERR] TypeError {e} в open_file_return_data")
            raise TypeError("Ошибка при открытии файла. Подробности в файле log_any_error.txt")
        except ValidateError as e:
            log_any_error(f"[ERR] ValidateError {e} в open_file_return_data")
            raise ValidateError(f"{e}")
        finally:
            try:
                wb._archive.close()
            except Exception as e:
                log_any_error(f"[ERR] Ошибка при закрытии файла. \n{e}")

    def dynamic_on_sheet(self, sheet, lst: list):
        """Обрабатывает лист с динамикой коек."""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        red_color = Color(rgb="FF0000")
        red_fill = PatternFill(patternType="solid", fgColor=red_color)

        for row in range(0, len(lst)):
            for column in range(0, len(lst[0])):
                if row == 0:
                    sheet.cell(row=row + 1, column=column + 1).value = lst[row][column]
                    sheet.cell(row=row + 1, column=column + 1).border = thin_border
                    sheet.cell(row=row + 1, column=column + 1).font = Font(bold=True)
                    continue
                sheet.cell(row=row + 1, column=column + 1).value = lst[row][column]
                sheet.cell(row=row + 1, column=column + 1).border = thin_border
                if column == 3 and int(lst[row][column]) < 0:
                    sheet.cell(row=row + 1, column=column + 1).fill = red_fill
                    sheet.cell(row=row + 1, column=column + 1).font = Font(
                        color="FFFFFF"
                    )
        for r in sheet.iter_rows(max_row=1):
            for col in r:
                col.alignment = Alignment(
                    vertical="center", horizontal="center", wrap_text=True
                )

        width_3 = 15
        sheet.column_dimensions["A"].width = 120
        sheet.column_dimensions["B"].width = width_3
        sheet.column_dimensions["C"].width = width_3
        sheet.column_dimensions["D"].width = width_3

    def fifty_on_sheet(self, sheet, lst: list):
        """Более 50(COUNT_DAYS) дней на лист эксель."""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in range(0, len(lst)):
            for column in range(0, len(lst[0])):
                if row == 0:
                    sheet.cell(row=row + 1, column=column + 1).value = lst[row][column]
                    sheet.cell(row=row + 1, column=column + 1).border = thin_border
                    sheet.cell(row=row + 1, column=column + 1).font = Font(bold=True)
                    continue
                sheet.cell(row=row + 1, column=column + 1).value = lst[row][column]
                sheet.cell(row=row + 1, column=column + 1).border = thin_border

        width_2 = 15
        sheet.auto_filter.ref = sheet.dimensions
        sheet.column_dimensions["A"].width = 120
        sheet.column_dimensions["B"].width = width_2
        sheet.column_dimensions["C"].width = width_2
        for r in sheet.iter_rows(max_row=1):
            for col in r:
                col.alignment = Alignment(
                    vertical="center", horizontal="center", wrap_text=True
                )

    def processing(
        self, data_for_bunks: dict, data_for_50: list
    ) -> Tuple[list, list, list, list]:
        """Обработка данных и запись в файл."""

        excel_bunks_kc = [
            [
                "Отделение",
                "Коечный фонд",
                "Кол-во госпитализированных на 9:00",
                "Свободных коек",
            ],
        ]
        excel_bunks_dc = [
            [
                "Отделение",
                "Коечный фонд",
                "Кол-во госпитализированных на 9:00",
                "Свободных коек",
            ],
        ]

        if not self.open_file_with_bunks():
            raise FileNotFoundError
        for department, count in data_for_bunks.items():
            try:
                free_bunks = BUNKS[department] - count
            except KeyError as e:
                log_any_error(
                    f'[ERR] Отделения {e} не оказалось в файле "Отделения и койки.xlsx"'
                )
                continue
            if (
                "ДС" in department
                or "днев" in department.lower()
                or "дн.стац." in department.lower()
                or "дн. стац." in department.lower()
            ):
                excel_bunks_dc.append(
                    [department, BUNKS[department], count, free_bunks]
                )
                continue
            excel_bunks_kc.append([department, BUNKS[department], count, free_bunks])

        excel_50_kc = [
            ["Отделение", "№ КВС", "Время пребывания в стационаре, в днях"],
        ]
        excel_50_dc = [
            ["Отделение", "№ КВС", "Время пребывания в стационаре, в днях"],
        ]
        for row in data_for_50:
            if (
                "ДС" in row[0]
                or "днев" in row[0].lower()
                or "дн.стац." in row[0].lower()
                or "дн. стац." in row[0].lower()
            ):
                excel_50_dc.append(row)
                continue
            excel_50_kc.append(row)
        return (excel_bunks_kc, excel_bunks_dc, excel_50_kc, excel_50_dc)

    def save_in_files(
        self,
        excel_bunks_kc: list,
        excel_bunks_dc: list,
        excel_50_kc: list,
        excel_50_dc: list,
    ):
        """Сохраняем данные на 4 листах."""
        wb = Workbook()
        wb.create_sheet(f"Более {COUNT_DAYS} дней")
        wb.create_sheet("Динамика коек ДС")
        wb.create_sheet(f"Более {COUNT_DAYS} дней ДС")

        sheet_bunks_kc = wb.active
        sheet_bunks_kc.title = "Динамика коек"
        self.dynamic_on_sheet(sheet_bunks_kc, excel_bunks_kc)
        sheet_bunks_dc = wb["Динамика коек ДС"]
        self.dynamic_on_sheet(sheet_bunks_dc, excel_bunks_dc)

        sheet_50_kc = wb[f"Более {COUNT_DAYS} дней"]
        self.fifty_on_sheet(sheet_50_kc, excel_50_kc)
        sheet_50_dc = wb[f"Более {COUNT_DAYS} дней ДС"]
        self.fifty_on_sheet(sheet_50_dc, excel_50_dc)

        wb.save(f"Койки и более {COUNT_DAYS} на {now()}.xlsx")

    @staticmethod
    def create_sample():
        """Создает файл образец."""
        sample = [
            ["Отделение из отчета", "Количество коек"],
            ["1025. Гинекологическое отделение", 50],
            ["1070. Терапевтическое отделение №1 (Коминтерна)", 60],
        ]
        wb_sample = Workbook()
        sheet_sample = wb_sample.active
        sheet_sample.title = "Пример заполнения"

        for record in sample:
            sheet_sample.append(record)
        sheet_sample.column_dimensions["A"].width = 80
        sheet_sample.column_dimensions["B"].width = 20
        wb_sample.save("Отделения и койки.xlsx")


def test():
    filepath = rf"{BASE_DIR}/Койки и более 50.xlsx"

    emk = BunkReport(filepath)
    data = emk.open_file_return_data()
    # # print(data)
    data_after = emk.processing(data[1], data[2])
    # emk.save_in_files(*data_after)


def test_1():
    # filepath = rf"{BASE_DIR}/files/койки 50.xlsx"
    filepath = rf"{BASE_DIR}/files/1696599179_hosp_EvnSection_List_pg.xlsx" 

    emk = BunkReport(filepath)
    data = emk.open_file_return_data()
    # # print(data)
    data_after = emk.processing(data[0], data[1])
    emk.save_in_files(*data_after)


if __name__ == "__main__":
    test_1()
