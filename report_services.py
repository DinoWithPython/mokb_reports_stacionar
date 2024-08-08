import datetime
from copy import deepcopy
import os
import sys
from typing import Any, Optional, Tuple, Union

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Color, Font, PatternFill
from openpyxl.styles.borders import Border, Side

from validators import (
    validate_column_with_data,
    validate_for_title,
    validate_numbers,
    validate_not_pdo,
    ValidateError)

if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
elif __file__:
    BASE_DIR = os.path.dirname(__file__)

EXPECTED_MIN_COLUMN_VALUES = 7
TITLE_VALUES = ["Наименование услуги", ]
LIS_SERVICERS = ["A08", "A09", "A12", "A26", "B03",]
HEADINGS = {
    0: "Отделение направления",
    1: "Врач",
    2: "Дата направления",
    3: "№ Направления",
    4: "КОД Услуги",
    5: "Наименование услуги",
    6: "ФИО пациента",
    7: "Дата рождения",
    8: "Дата выполнения услуги",
}


def now() -> str:
    return datetime.datetime.now().strftime("%d.%m %H_%M_%S")


def log_any_error(any_text):
    """Функция для записи ошибок в файл."""
    with open(rf"{BASE_DIR}\log_any_error.txt", "a+", encoding="UTF8") as log:
        text = str(any_text)
        print(f"[{now()}] {text}", file=log)


def check_index(lst: list, column_name: str, class_name: str = None):
    """Получает название столбца и возвращает индекс."""
    try:
        upper_list = [element.upper() if isinstance(element, str) else element for element in lst]
        index_el = upper_list.index(column_name.upper())
        return index_el
    except ValueError:
        if class_name:
            log_any_error(f'Не найден столбец "{column_name}" в классе {class_name}')
            return
        log_any_error(f'Не найден столбец "{column_name}"')


def date_conversion(date: Union[str, datetime.datetime]) -> str:
    """Преобразует дату к нужному формату."""
    if isinstance(date, datetime.datetime):
        return date.strftime("%d.%m.%Y")
    if isinstance(date, str):
        raise ValueError("Попытка преобразовать дату из строки, а не datetime. Автор не был к такому готов. Нужна переработка скрипта. :)")


def check_convert_type(variable: Any, type_variable: Any, name_variable: str):
    """Проверяет возможность преобразования в переданный тип."""
    try:
        type_variable(variable)
    except ValueError as e:
        text_err = (
            f"Ошибка в типе переменной. Типом '{variable}' является '{type(variable)}'."
            f"\nОжидается преобразование {name_variable} в '{type_variable}'!")
        log_any_error(text_err, e)
        raise ValueError(text_err)



class ServicesReport:
    """Класс для создания отчетов по услугам. Принимает путь для файла."""
    period = set()

    def __init__(self, filepath: str):
        self.filepath = filepath
        self.department = None
        self.doctor = None
        self.date_direct = None
        self.number_direct = None
        self.code_service = None
        self.name_service = None
        self.fio_patient = None
        self.brthday_patient = None
        self.date_complete_service = None

    def __str__(self):
        return "ServicesReport"

    def validate_data_from_file(self, works_sheet, lst: list):
        """Валидация данных из файла."""
        title_excel = []
        for row in works_sheet.iter_rows(values_only=True):
            if validate_column_with_data(row, EXPECTED_MIN_COLUMN_VALUES):
                continue
            if validate_numbers(row):
                continue
            if len(title_excel) == 0:
                if validate_for_title(TITLE_VALUES, row):
                    title_excel.extend(row)
                    self.department = check_index(title_excel, HEADINGS[0], self)
                    self.doctor = check_index(title_excel, HEADINGS[1], self)
                    self.date_direct = check_index(title_excel, HEADINGS[2], self)
                    self.number_direct = check_index(title_excel, HEADINGS[3], self)
                    self.code_service = check_index(title_excel, HEADINGS[4], self)
                    self.name_service = check_index(title_excel, HEADINGS[5], self)
                    self.fio_patient = check_index(title_excel, HEADINGS[6], self)
                    self.brthday_patient = check_index(title_excel, HEADINGS[7], self)
                    self.date_complete_service = check_index(title_excel, HEADINGS[8], self)
                    if None in self.__dict__.values():
                        raise ValueError("Отсутвует необходимый столбец. Подробности в файле log_any_error.txt")
                    continue
            lst.append(row)

        if len(title_excel) == 0:
            log_any_error(f"Заголовки не были найдены! Ни в одной строке не было: \n{TITLE_VALUES}\n")
            raise ValueError("В файле отсутствуют заголовки! Подробности в файле log_any_error.txt")

    def open_file_return_data(self) -> Tuple[list, list, dict]:
        """Открыли файл и вернули истину и данные, либо ложь и ошибку."""
        try:
            wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
            ws = wb.active

            inst_from_excel = []
            lis_from_excel = []
            data_from_excel = []
            data_for_svod = {
                'intrumental': {},
                'laboratory': {}
            }
            sheets = wb.sheetnames
            if len(sheets) == 1:
                ws = wb.active
                self.validate_data_from_file(ws, data_from_excel)
            else:
                for sheet in sheets:
                    ws = wb[sheet]
                    self.validate_data_from_file(ws, data_from_excel)
            for row in data_from_excel:
                if row[self.date_complete_service] is not None:
                    if row[self.code_service][:3].upper() in LIS_SERVICERS:
                        data_for_svod['laboratory'].setdefault(row[self.department], [0, 0])
                        data_for_svod['laboratory'][row[self.department]][0] += 1
                        data_for_svod['laboratory'][row[self.department]][1] += 1
                        continue
                    data_for_svod['intrumental'].setdefault(row[self.department], [0, 0])
                    data_for_svod['intrumental'][row[self.department]][0] += 1
                    data_for_svod['intrumental'][row[self.department]][1] += 1
                    continue
                self.period.add(row[self.date_direct])
                if row[self.code_service][:3].upper() in LIS_SERVICERS:
                    data_for_svod['laboratory'].setdefault(row[self.department], [0, 0])
                    data_for_svod['laboratory'][row[self.department]][0] += 1
                    lis_from_excel.append([
                        row[self.department],
                        row[self.doctor],
                        date_conversion(row[self.date_direct]),
                        row[self.number_direct],
                        row[self.code_service],
                        row[self.name_service],
                        row[self.fio_patient],
                        date_conversion(row[self.brthday_patient]),
                        row[self.date_complete_service],
                    ])
                    continue
                data_for_svod['intrumental'].setdefault(row[self.department], [0, 0])
                data_for_svod['intrumental'][row[self.department]][0] += 1
                inst_from_excel.append([
                    row[self.department],
                    row[self.doctor],
                    date_conversion(row[self.date_direct]),
                    row[self.number_direct],
                    row[self.code_service],
                    row[self.name_service],
                    row[self.fio_patient],
                    date_conversion(row[self.brthday_patient]),
                    row[self.date_complete_service],
                    ])
            if len(inst_from_excel) == 0:
                log_any_error("Нет данных инструментальных исследований!")
            if len(lis_from_excel) == 0:
                log_any_error("Нет данных лаборабортных исследований!")
            

            # Добавляем процент и итог
            data_for_svod['intrumental']['Итого'] = [0, 0, 0]
            for department, values in data_for_svod['intrumental'].items():
                data_for_svod['intrumental'][department].append(
                    values[1] / values[0] if values[0] != 0 else 0)
                if department != 'Итого':
                    data_for_svod['intrumental']['Итого'][0] += data_for_svod[
                        'intrumental'][department][0]
                    data_for_svod['intrumental']['Итого'][1] += data_for_svod[
                        'intrumental'][department][1]
            data_for_svod['intrumental']['Итого'][2] = data_for_svod[
                'intrumental']['Итого'][1] / data_for_svod['intrumental']['Итого'][0]

            data_for_svod['laboratory']['Итого'] = [0, 0, 0]
            for department, values in data_for_svod['laboratory'].items():
                data_for_svod['laboratory'][department].append(
                    values[1] / values[0] if values[0] != 0 else 0)
                if department != 'Итого':
                    data_for_svod['laboratory']['Итого'][0] += data_for_svod[
                            'laboratory'][department][0]
                    data_for_svod['laboratory']['Итого'][1] += data_for_svod[
                            'laboratory'][department][1]
            data_for_svod['laboratory']['Итого'][2] = data_for_svod[
                'laboratory']['Итого'][1] / data_for_svod['laboratory']['Итого'][0]
            return inst_from_excel, lis_from_excel, data_for_svod
        except TypeError as err:
            raise ValidateError(err) from err
        except Exception as err:
            raise ValidateError(err) from err
        finally:
            try:
                wb._archive.close()
            except Exception as e:
                log_any_error(f"[ERR] Ошибка при закрытии файла. \n{e}")

    def data_at_sheet_payroll(self, sheet, lst: list):
        """Добавляет данные и оформление для списочного листа."""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        red_color = Color(rgb="FF0000")
        red_fill = PatternFill(patternType="solid", fgColor=red_color)
        if len(self.period) > 1:
            date_min = date_conversion(min(self.period))
            date_max = date_conversion(max(self.period))
        else:
            date_min = date_max = date_conversion(self.period)
        
        sheet.cell(row=1, column=1).value = (
            "Список пациентов, которым выданы направления на услуги "
            f"{date_min}-{date_max}"
        )
        for c in range(1, len(HEADINGS)+1):
            sheet.cell(row=2, column=c).value = HEADINGS[c-1]
            sheet.cell(row=2, column=c).border = thin_border
            sheet.cell(row=2, column=c).font = Font(bold=True)
            sheet.cell(row=2, column=c).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
        for r in range(2, len(lst)+2):
            for c in range(0, len(HEADINGS)):
                sheet.cell(row=r + 1, column=c + 1).value = lst[r-2][c]
                sheet.cell(row=r + 1, column=c + 1).border = thin_border
                if c == len(HEADINGS)-1:
                    sheet.cell(row=r + 1, column=c + 1).fill= red_fill
            # sheet.merge_cells(
            #     start_row=1, start_column=column+1, end_row=1, end_column=column+2
            # )
        sheet.auto_filter.ref = "A2:I2"

        # Задаю ширину столбцев свода
        sheet.column_dimensions["A"].width = 25
        sheet.column_dimensions["B"].width = 15
        sheet.column_dimensions["C"].width = 15
        sheet.column_dimensions["D"].width = 10
        sheet.column_dimensions["E"].width = 15
        sheet.column_dimensions["F"].width = 15
        sheet.column_dimensions["G"].width = 25
        sheet.column_dimensions["H"].width = 15


    def data_at_sheet_svod(self, data: dict, sheet, type_service: str):
        """Добавляет данные на лист свода."""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        if len(self.period) > 1:
            date_min = date_conversion(min(self.period))
            date_max = date_conversion(max(self.period))
        else:
            date_min = date_max = date_conversion(self.period)

        sheet.append([
            f"Список пациентов, которым выданы направления на {type_service} услуги "
            f"{date_min}-{date_max}"
            ])

        sheet.append([
            "Подразделение",
            f"Количество оформленных направлений на {type_service} услуги",
            f"Количество выполненных направлений на {type_service} услуги",
            "Процент оформленных направлений на выполненные"
        ])
        for department, values in data.items():
            sheet.append([department, values[0], values[1], values[2]])

        sheet.auto_filter.ref = "A2:D2"
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        my_color = Color(rgb="AFEEEE")
        my_fill = PatternFill(patternType="solid", fgColor=my_color)
        for r in sheet.iter_rows(max_row=2):
            for column in r:
                column.border = thin_border
                column.fill = my_fill
                column.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

        red_color = Color(rgb="FF0000")
        red_fill = PatternFill(patternType="solid", fgColor=red_color)
        green_color = Color(rgb="32CD32")
        green_fill = PatternFill(patternType="solid", fgColor=green_color)
        for r in sheet.iter_rows(min_row=3):
            for column in range(len(r)):
                r[column].border = thin_border
                r[column].alignment = Alignment(horizontal="center", vertical="center")
                if column == 1 or column == 2:
                    if r[column].value == 0:
                        r[column].value = None
                if column == 3:
                    r[column].fill = (
                        red_fill if float(r[column].value) < 0.80 else green_fill
                    )
                    r[column].number_format = "0%"

        sheet.column_dimensions["A"].width = 100
        columns_for_dimensions = ["B", "C", "D"]
        for letter in columns_for_dimensions:
            sheet.column_dimensions[letter].width = 15


    def save_files(self, data_inst: list, data_lis: list, data_svod: dict):
        """Функция для сохранения списочных файлов."""
        wb_inst = Workbook()
        sheet_inst= wb_inst.active
        sheet_inst.title = "Инструм"
        self.data_at_sheet_payroll(sheet_inst, data_inst)
        wb_inst.create_sheet("Свод инструм")
        sheet_isnt_svod = wb_inst["Свод инструм"]
        self.data_at_sheet_svod(data_svod['intrumental'], sheet_isnt_svod, "инструм.")
        wb_inst.save(f"Инструм. услуги на {now()}.xlsx")

        wb_lis = Workbook()
        sheet_lis= wb_lis.active
        sheet_lis.title = "ЛИС"
        self.data_at_sheet_payroll(sheet_lis, data_lis)
        wb_lis.create_sheet("Свод ЛИС")
        sheet_lis_svod = wb_lis["Свод ЛИС"]
        self.data_at_sheet_svod(data_svod['laboratory'], sheet_lis_svod, "лабор.")
        wb_lis.save(f"ЛИС услуги на {now()}.xlsx")



if __name__ == "__main__":
    def test_1():
        # filepath = rf"{BASE_DIR}/files/Список пациентов, которым выданы направления на услуги.xlsx"
        filepath = rf"{BASE_DIR}/files/1696599422_pan_SpisokDirection_Usluga_pg.xlsx"

        report = ServicesReport(filepath)
        inst, lis, svod = report.open_file_return_data()
        report.save_files(inst, lis, svod)

    test_1()


