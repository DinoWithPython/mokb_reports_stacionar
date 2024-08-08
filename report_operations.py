import datetime
import os
import sys
from typing import Optional, Tuple, Union

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Color, Font, NamedStyle, PatternFill
from openpyxl.styles.borders import Border, Side

from validators import (
    validate_column_with_data,
    validate_for_title,
    validate_numbers)

if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
elif __file__:
    BASE_DIR = os.path.dirname(__file__)

EXPECTED_MIN_COLUMN_VALUES = 7
TITLE_VALUES_LEFT = ["№ КВС", ]
TITLE_VALUES_RIGHT = ["Дата проведения операции", ]
SERVICE_CODE = "a16."
HEADINGS = {
    0: "№ КВС",
    1: "ФИО пациента",
    2: "Наименова-\nние\nотделения",
    3: "Дата поступле-ния",
    4: "Код операции",
    5: "Наличие\nпротокола\nоперации",
    6: "Дата проведения операции",
    7: "Количество\nопераций",
    8: "Наименова-ние операции",
}

def now():
    return datetime.datetime.now().strftime("%d.%m %H_%M_%S")


def log_any_error(any_text):
    """Функция для записи ошибок в файл."""
    log = open(rf"{BASE_DIR}\log_any_error.txt", "a+")
    text = str(any_text)
    print(f"[{now()}] {text}", file=log)
    log.close()


def check_index(lst: list, column_name: str, class_name: str = None):
    """Получает название столбца и возвращает индекс."""
    try:
        upper_list = [element.upper() if isinstance(element, str) else element for element in lst]
        index_el = upper_list.index(column_name.upper())
        return index_el
    except ValueError:
        if class_name:
            log_any_error(f'Не найден столбец "{column_name}" в классе {class_name}')
            return
        log_any_error(f'Не найден столбец "{column_name}"')


class OperationReport:
    """Класс для создания отчетов по операия. Принимает путь для файла."""

    def __init__(self, filepath: str):
        self.filepath = filepath
        self.kvs_number = None
        self.last_name = None
        self.department = None
        self.date_host_in = None
        self.code_operation = None
        self.protocol_operation = None
        self.date_operation = None
        self.count_operations = None
        self.operation_column = None
        self.operaion_name = None

    def __str__(self):
        return "OperationReport"

    def validate_data_from_file(self, works_sheet, for_data: list, only_a16: bool = True):
        """Валидация данных из файла."""
        title_excel_left = []
        title_excel_right = []
        for row in works_sheet.iter_rows(values_only=True):
            if validate_column_with_data(row, EXPECTED_MIN_COLUMN_VALUES):
                continue
            if len(title_excel_left) == 0:
                if validate_for_title(TITLE_VALUES_LEFT, row):
                    title_excel_left.extend(row)
                    self.kvs_number = check_index(title_excel_left, HEADINGS[0], self)
                    self.last_name = check_index(title_excel_left, HEADINGS[1], self)
                    continue
                continue
            if len(title_excel_right) == 0:
                if validate_for_title(TITLE_VALUES_RIGHT, row):
                    title_excel_right.extend(row)
                    self.department = check_index(title_excel_right, HEADINGS[2], self)
                    self.date_host_in = check_index(title_excel_right, HEADINGS[3], self)
                    self.code_operation = check_index(title_excel_right, HEADINGS[4], self)
                    self.protocol_operation = check_index(title_excel_right, HEADINGS[5], self)
                    self.date_operation = check_index(title_excel_right, HEADINGS[6], self)
                    self.count_operations = check_index(title_excel_right, HEADINGS[7], self)
                    self.operaion_name = check_index(title_excel_right, HEADINGS[8], self)
                    self.operation_column = (self.date_operation, self.count_operations)
                    continue
                continue
            if validate_numbers(row):
                continue
            if None in self.__dict__.values():
                raise ValueError("Отсутвует необходимый столбец. Подробности в файле log_any_error.txt")
            if row[self.code_operation] is not None:
                if 'вентиляц' in row[self.operaion_name].lower():
                    continue
                if only_a16:
                    if row[self.code_operation].lower().startswith(SERVICE_CODE.lower()):
                        last_name = row[self.last_name].split()[0]
                        lst = [
                            row[self.kvs_number],
                            last_name,
                            row[self.department],
                            row[self.date_host_in],
                        ]
                        lst.extend(row[self.operation_column[0]:self.protocol_operation])
                        prt_operation = "Да" if row[self.protocol_operation] != 0 else "Нет"
                        lst.append(prt_operation)
                        lst.extend(row[self.protocol_operation+1:self.operation_column[-1]+1])
                        for_data.append(lst)
                    continue
                lst = [
                    row[self.kvs_number],
                    row[self.department],
                    row[self.date_host_in],
                ]
                lst.extend(row[self.operation_column[0]:self.protocol_operation])
                prt_operation = "Да" if row[self.protocol_operation] != 0 else "Нет"
                lst.append(prt_operation)
                lst.extend(row[self.protocol_operation+1:self.operation_column[-1]+1])
                for_data.append(lst)

        if len(title_excel_left) == 0 or len(title_excel_right) == 0:
            log_any_error(f"Заголовки не были найдены! Ни в одной строке не было: \n{TITLE_VALUES_LEFT}\n\n{TITLE_VALUES_RIGHT}\n")
            raise ValueError("В файле отсутствуют заголовки! Подробности в файле log_any_error.txt")
        if len(for_data) == 0:
            log_any_error("Файл с данными пуст!")
            raise ValueError("Файл с данными пуст! Проверьте, что выбрали нужный файл.")

    def open_file_return_data(self, only_a16: bool) -> list:
        """Открыли файл и вернули истину и данные, либо ложь и ошибку."""
        try:
            wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)

            data = [
                (
                    "Номер карты",
                    "Фамилия",
                    "Профильное отделение",
                    "Дата поступления",
                    "Дата проведения операции",
                    "Время начала операции",
                    "Тип операции плановая/экстренная",
                    "Код операции",
                    "Наименование операции",
                    "Наличие предоперационного эпикриза",
                    "Наличие протокола операции",
                    "Осложнения",
                    "ФИО хирурга",
                    "Количество операций",
                ),
            ]

            sheets = wb.sheetnames
            if len(sheets) == 1:
                ws = wb.active
                self.validate_data_from_file(ws, data, only_a16)
            else:
                for sheet in sheets:
                    ws = wb[sheet]
                    self.validate_data_from_file(ws, data, only_a16)
            return data
        except TypeError as e:
            log_any_error(f"[ERR] {e}")
            return TypeError("Не удалось открыть файл!")
        finally:
            try:
                wb._archive.close()
            except Exception as e:
                log_any_error(f"[ERR] Ошибка при закрытии файла. \n{e}")

    def processing_and_save(self, data):
        """Функция для сохранения отчета."""
        wb = Workbook()
        sheet_operations = wb.active
        sheet_operations.title = "Операции"
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        full_date = NamedStyle(name="full_date", number_format="DD/MM/YYYY HH:MM")
        only_time = NamedStyle(name="only_time", number_format="HH:MM")
        red_color = Color(rgb="FF0000")

        for row in range(0, len(data)):
            for column in range(0, len(data[0])):
                if row == 0:
                    sheet_operations.cell(row=row + 1, column=column + 1).value = data[
                        row
                    ][column]
                    sheet_operations.cell(
                        row=row + 1, column=column + 1
                    ).border = thin_border
                    continue
                if column == 2 or column == 3:
                    sheet_operations.cell(
                        row=row + 1, column=column + 1
                    ).style = full_date
                if column == 4:
                    sheet_operations.cell(
                        row=row + 1, column=column + 1
                    ).style = only_time
                sheet_operations.cell(row=row + 1, column=column + 1).value = data[row][
                    column
                ]
                sheet_operations.cell(
                    row=row + 1, column=column + 1
                ).border = thin_border

        red_fill = PatternFill(patternType="solid", fgColor=red_color)
        for r in sheet_operations.iter_rows(min_row=2):
            if r[data[0].index("Наличие протокола операции")].value == "Нет":
                r[data[0].index("Наличие протокола операции")].fill = red_fill

        sheet_operations.column_dimensions["A"].width = 15
        sheet_operations.column_dimensions["B"].width = 20
        sheet_operations.column_dimensions["C"].width = 50
        sheet_operations.column_dimensions["D"].width = 20
        sheet_operations.column_dimensions["E"].width = 20
        sheet_operations.column_dimensions["F"].width = 14
        sheet_operations.column_dimensions["G"].width = 15
        sheet_operations.column_dimensions["H"].width = 23
        sheet_operations.column_dimensions["I"].width = 23
        sheet_operations.auto_filter.ref = sheet_operations.dimensions

        for r in sheet_operations.iter_rows(max_row=1):
            for col in r:
                col.font = Font(bold=True)
                col.alignment = Alignment(
                    vertical="center", horizontal="center", wrap_text=True
                )
        wb.save(f"Операции на {now()}.xlsx")


def test():
    filepath = rf"{BASE_DIR}/files/операции.xlsx"

    emk = OperationReport(filepath)
    data = emk.open_file_return_data(True)
    emk.processing_and_save(data)


if __name__ == "__main__":
    test()
