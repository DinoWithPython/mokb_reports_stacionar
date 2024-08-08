import datetime
import os
import sys
from typing import Optional, Tuple, Union

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side

from validators import (
    validate_column_with_data,
    validate_for_title,
    validate_numbers)


if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
elif __file__:
    BASE_DIR = os.path.dirname(__file__)

TITLE_VALUES_UP = ["Номер карты", ]
TITLE_VALUES_DOWN = ["Отделение", ]
EXPECTED_MIN_COLUMN_VALUES = 7
HEADINGS = {
    0: "Номер карты",
    1: "Адрес проживания/ регистрации",
    2: "Телефон",
    3: "Отделение",
}

TITLE = [["Отделение", "№ КВС"]]


def now():
    return datetime.datetime.now().strftime("%d.%m %H_%M_%S")


def log_any_error(any_text):
    """Функция для записи ошибок в файл."""
    log = open(rf"{BASE_DIR}\log_any_error.txt", "a+")
    text = str(any_text)
    print(f"[{now()}] {text}", file=log)
    log.close()


def check_index(lst: list, column_name: str, class_name: str = None):
    """Получает название столбца и возвращает индекс."""
    try:
        upper_list = [element.upper() if isinstance(element, str) else element for element in lst]
        index_el = upper_list.index(column_name.upper())
        return index_el
    except ValueError:
        if class_name:
            log_any_error(f'Не найден столбец "{column_name}" в классе {class_name}')
            return
        log_any_error(f'Не найден столбец "{column_name}"')


class PhoneReport:
    """Класс для создания отчетов по ЭМК. Принимает путь для файла."""

    def __init__(self, filepath: str):
        self.filepath = filepath
        self.number_cart = None
        self.adress = None
        self.phone = None
        self.department = None

    def validate_data_from_file(self, works_sheet, for_data_phone: list, for_data_adress: list):
        """Валидация данных из файла."""
        title_excel_up = []
        title_excel_down = []
        for row in works_sheet.iter_rows(values_only=True):
            if validate_column_with_data(row, EXPECTED_MIN_COLUMN_VALUES):
                continue
            if len(title_excel_up) == 0:
                if validate_for_title(TITLE_VALUES_UP, row):
                    title_excel_up.extend(row)
                    self.number_cart = check_index(title_excel_up, HEADINGS[0], self)
                    self.adress = check_index(title_excel_up, HEADINGS[1], self)
                    self.phone = check_index(title_excel_up, HEADINGS[2], self)
                    continue
                continue
            if len(title_excel_down) == 0:
                if validate_for_title(TITLE_VALUES_DOWN, row):
                    title_excel_down.extend(row)
                    self.department = check_index(title_excel_down, HEADINGS[3], self)
                    continue
                continue
            if validate_numbers(row):
                continue
            if None in self.__dict__.values():
                raise ValueError("Отсутвует необходимый столбец. Подробности в файле log_any_error.txt")
            if len(row[self.adress] if row[self.adress] is not None else []) < 10 :
                for_data_adress.append([row[self.department], row[self.number_cart]])
            if row[self.phone] is None:
                for_data_phone.append([row[self.department], row[self.number_cart]])

        if len(title_excel_up) == 0 or len(title_excel_down) == 0:
            log_any_error(f"Заголовки не были найдены! Ни в одной строке не было: \n{title_excel_up}\n\n{title_excel_down}\n")
            raise ValueError("В файле отсутствуют заголовки! Подробности в файле log_any_error.txt")

    def open_file_return_data(self) -> Union[list, list]:
        """Открыли файл и вернули истину и данные, либо ложь и ошибку."""
        try:
            wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
            ws = wb.active

            data_phone = []
            data_adress = []
            sheets = wb.sheetnames
            if len(sheets) == 1:
                ws = wb.active
                self.validate_data_from_file(ws, data_phone, data_adress)
            else:
                for sheet in sheets:
                    ws = wb[sheet]
                    self.validate_data_from_file(ws, data_phone, data_adress)
            return data_phone, data_adress
        except TypeError as e:
            log_any_error(f"[ERR] {e}")
            raise TypeError("Ошибка при открытии файла.")
        finally:
            try:
                wb._archive.close()
            except Exception as e:
                log_any_error(f"[ERR] Ошибка при закрытии файла. \n{e}")

    def data_on_sheet(self, sheet, data: list):
        """Принимает лист и данные для формирования."""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in range(0, len(data)):
            for column in range(0, len(data[0])):
                if row == 0:
                    sheet.cell(row=row + 1, column=column + 1).value = TITLE[row][
                        column
                    ]
                    sheet.cell(row=row + 1, column=column + 1).border = thin_border
                    continue
                sheet.cell(row=row + 1, column=column + 1).value = data[row][column]
                sheet.cell(row=row + 1, column=column + 1).border = thin_border

        sheet.column_dimensions["A"].width = 80
        sheet.column_dimensions["B"].width = 15
        sheet.auto_filter.ref = sheet.dimensions
        for r in sheet.iter_rows(max_row=1):
            for col in r:
                col.font = Font(bold=True)
                col.alignment = Alignment(
                    vertical="center", horizontal="center", wrap_text=True
                )

    def processing_and_save(self, data_phone, data_adress):
        """Формирует отчеты и сохраняет в эксель."""
        wb = Workbook()
        wb.create_sheet("Без телефона ДС")
        wb.create_sheet("Без адреса")
        wb.create_sheet("Без адреса ДС")

        sheet_phone_kc = wb.active
        sheet_phone_kc.title = "Без телефона"
        sheet_phone_dc = wb["Без телефона ДС"]
        phone_kc = []
        phone_dc = []

        DEPARTMENT = 0
        for record in data_phone:
            if (
                "ДС" in record[DEPARTMENT]
                or "днев" in record[DEPARTMENT].lower()
                or "дн.стац." in record[DEPARTMENT].lower()
                or "дн. стац." in record[DEPARTMENT].lower()
            ):
                phone_dc.append(record)
                continue
            phone_kc.append(record)

        phone_kc.sort()
        phone_dc.sort()
        self.data_on_sheet(sheet_phone_kc, phone_kc)
        self.data_on_sheet(sheet_phone_dc, phone_dc)

        sheet_adress_kc = wb["Без адреса"]
        sheet_adress_dc = wb["Без адреса ДС"]
        adress_kc = []
        adress_dc = []
        for record in data_adress:
            if (
                "ДС" in record[DEPARTMENT]
                or "днев" in record[DEPARTMENT].lower()
                or "дн.стац." in record[DEPARTMENT].lower()
                or "дн. стац." in record[DEPARTMENT].lower()
            ):
                adress_dc.append(record)
                continue
            adress_kc.append(record)
        adress_kc.sort()
        adress_dc.sort()
        self.data_on_sheet(sheet_adress_kc, adress_kc)
        self.data_on_sheet(sheet_adress_dc, adress_dc)

        wb.save(f"Телефоны и адреса на {now()}.xlsx")


def test():
    filepath = rf"{BASE_DIR}/files/Список поступивших пациентов по дате и времени.xlsx"
    # filepath = rf"{BASE_DIR}/files/1696599142_han_evnps_timelist_new_pg.xlsx"

    emk = PhoneReport(filepath)
    data = emk.open_file_return_data()
    emk.processing_and_save(data[0], data[1])


if __name__ == "__main__":
    test()
